# %% [markdown]
# This is for training testing and validating, Shared weights stereo ConvNeXt V2 Model @ 224 with custom 6ch input head
# for use in 6D single object prediction

# %% [markdown]
# IMPORTS

# %%
import os
# import json
import csv
import time
import torch
import timm
import torch.nn as nn
from torch.nn.utils import prune
import torch.nn.functional as F
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader
from torchvision import transforms
from PIL import Image
import numpy as np
import pandas as pd
import datetime
from tqdm import tqdm
from typing import Tuple
from torch.amp import GradScaler, autocast
from torch.utils.tensorboard import SummaryWriter
from openpyxl import Workbook, load_workbook

# %% [markdown]
# DCLRATIONS

# %%

# with open("GlobVar.json", "r") as file:
#     gv = json.load(file)

# mod_id = gv['mod_id']
# MOD_ID = mod_id
BATCH_ID = 5
MOD_ID = 3
BATCH_SIZE = 32
NUM_EPOCHS = 30
LEARNING_RATE = 1e-4
TRANS_WEIGHT = 1.5
ROTATION_WEIGHT = 1.0
ANGULAR_WEIGHT = 0.1
PATIENCE = 3
IMG_SIZE = 224

# %%
BASE_DIR = os.path.expanduser("~/SKRIPSI/SCRIPTS")
DATASET_DIR = os.path.join(BASE_DIR, f"dataset/batch{BATCH_ID}")
MODEL_SAVE_PATH = os.path.join(BASE_DIR, f"model/S6ch-ConvNeXt6DP{BATCH_ID}.{MOD_ID}.pth")
BEST_MODEL_PATH = os.path.join(BASE_DIR, f"model/BEST-S6ch-ConvNeXt6DP{BATCH_ID}.{MOD_ID}.pth")
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# %% [markdown]
# DATASETCLASS

# %%
class Stereo6ChPoseDataset(Dataset):
    def __init__(self, csv_file, root_dir, transform=None):
        self.annotations = pd.read_csv(csv_file)
        self.root_dir = root_dir
        self.transform = transform

    def __len__(self):
        return len(self.annotations)

    def __getitem__(self, idx):
        img_path = os.path.join(self.root_dir, self.annotations.iloc[idx, 0])
        image = Image.open(img_path).convert("RGB")
        width, height = image.size
        mid = width // 2

        imageL = image.crop((0, 0, mid, height))
        imageR = image.crop((mid, 0, width, height))

        if self.transform:
            imageL = self.transform(imageL)  # (3, H, W)
            imageR = self.transform(imageR)  # (3, H, W)

        # Stack to make (6, H, W)
        image6ch = torch.cat([imageL, imageR], dim=0)

        label = torch.tensor(self.annotations.iloc[idx, 1:].astype(np.float32).values)
        return image6ch, label

# %% [markdown]
# Conversions loss functions rmse yadaydadaydada

# %%
def rotation_error(R_pred, R_gt):
    """Compute angular error in degrees between rotation matrices."""
    R_diff = torch.bmm(R_pred.transpose(1, 2), R_gt)
    trace = torch.diagonal(R_diff, dim1=1, dim2=2).sum(dim=1)
    eps = 1e-6
    angle_rad = torch.acos(torch.clamp((trace - 1) / 2, min=-1 + eps, max=1 - eps))
    return torch.rad2deg(angle_rad)


def geodesic_loss(R_pred, R_gt):
    R_diff = torch.bmm(R_pred.transpose(1, 2), R_gt)
    trace = torch.diagonal(R_diff, dim1=1, dim2=2).sum(dim=1)
    eps = 1e-6
    angle = torch.acos(torch.clamp((trace - 1) / 2, -1 + eps, 1 - eps))
    return angle.mean()

def compute_rotation_matrix_from_ortho6d(poses_6d):
    """Convert 6D rotation representation to 3x3 rotation matrices."""
    x_raw = poses_6d[:, 0:3]
    y_raw = poses_6d[:, 3:6]

    x = F.normalize(x_raw, dim=1)
    z = F.normalize(torch.cross(x, y_raw, dim=1), dim=1)
    y = torch.cross(z, x, dim=1)

    rot = torch.stack((x, y, z), dim=-1)  # Shape: [B, 3, 3]
    return rot

def combined_loss(output, target, trans_w=1.0, rot_w=1.0, ang_w=0.1):
    pred_trans = output[:, :3]
    gt_trans = target[:, :3]
    pred_rot_6d = output[:, 3:9]
    gt_rot_6d = target[:, 3:9]

    pred_rot = compute_rotation_matrix_from_ortho6d(pred_rot_6d)
    gt_rot = compute_rotation_matrix_from_ortho6d(gt_rot_6d)

    loss_trans = F.mse_loss(pred_trans, gt_trans)
    loss_rot = geodesic_loss(pred_rot, gt_rot)
    return trans_w * loss_trans + rot_w * loss_rot


def rotation_error_deg_from_6d(pred_6d, gt_6d):
    # Ensure same dtype (float32)
    pred_6d = pred_6d.float()
    gt_6d = gt_6d.float()

    R_pred = compute_rotation_matrix_from_ortho6d(pred_6d)
    R_gt = compute_rotation_matrix_from_ortho6d(gt_6d)

    R_diff = torch.bmm(R_pred.transpose(1, 2), R_gt)
    trace = R_diff[:, 0, 0] + R_diff[:, 1, 1] + R_diff[:, 2, 2]
    cos_theta = (trace - 1) / 2
    cos_theta = torch.clamp(cos_theta, -1.0, 1.0)

    theta = torch.acos(cos_theta)
    return torch.rad2deg(theta.mean())


def compute_errors(outputs, labels):
    outputs = outputs.float()
    labels = labels.float()

    pred_trans = outputs[:, :3]
    pred_rot_6d = outputs[:, 3:9]

    gt_trans = labels[:, :3]
    gt_rot_6d = labels[:, 3:9]

    trans_rmse = torch.sqrt(F.mse_loss(pred_trans, gt_trans))
    rot_rmse = rotation_error_deg_from_6d(pred_rot_6d, gt_rot_6d)

    return trans_rmse.item(), rot_rmse.item()


def calculate_translation_rmse(preds, gts):
    trans_rmse = np.sqrt(np.mean(np.sum((preds[:, :3] - gts[:, :3])**2, axis=1)))
    return trans_rmse * 100  # Convert m → cm

def translation_accuracy_percentage(rmse_cm, range_cm):
    return max(0.0, 100.0 * (1 - rmse_cm / range_cm))

# %% [markdown]
# Transform & Dataloader

# %%
def get_transform():
    return transforms.Compose([
        transforms.Resize((IMG_SIZE, IMG_SIZE)),
        transforms.ToTensor(),
        transforms.Normalize(
            mean=[0.485, 0.456, 0.406],  # ImageNet mean
            std=[0.229, 0.224, 0.225]    # ImageNet std
        )
    ])

def get_dataloader(split):
    base_dir = os.path.join(DATASET_DIR, split)
    csv_path = os.path.join(base_dir, "labels.csv")
    images_dir = os.path.join(base_dir, "images")  # Updated to point to the actual image folder
    dataset = Stereo6ChPoseDataset(csv_path, images_dir, transform=get_transform())
    return DataLoader(dataset, batch_size=BATCH_SIZE, shuffle=(split == "train"))

# %%
def get_dataset_stats(loader):
    translations = []
    for _, labels in loader:
        translations.append(labels[:, :3])
    trans = torch.cat(translations, dim=0)
    return {
        'min': trans.min(dim=0).values,
        'max': trans.max(dim=0).values,
        'mean': trans.mean(dim=0),
        'std': trans.std(dim=0)
    }

# %% [markdown]
# Loading...

# %%
train_loader = get_dataloader("train")
val_loader = get_dataloader("val")
test_loader = get_dataloader("test")

# %% [markdown]
# Model Definitions

# %%
class ConvNeXt6DP6ch(nn.Module):
    def __init__(self):
        super(ConvNeXt6DP6ch, self).__init__()

        self.backbone = timm.create_model(
            "convnextv2_nano.fcmae_ft_in22k_in1k",
            pretrained=True,
            features_only=False
        )
        self.backbone.head = nn.Identity()  # Remove classifier
        self._inflate_input_channels()      # Apply patch here 👈

        self.pool = nn.AdaptiveAvgPool2d((1, 1))  # Global Average Pooling
        self.flatten = nn.Flatten()

        self.head = nn.Sequential(
            nn.Linear(640, 512),  # Output of convnextv2_nano is 640
            nn.ReLU(),
            nn.Linear(512, 9),
            # nn.ReLU(),
            # nn.Linear(128, 9)
        )

    def _inflate_input_channels(self):
        old_conv = self.backbone.stem[0]  # Get the first Conv2d layer

        new_conv = nn.Conv2d(
            in_channels=6,  # Accept stereo (6-channel)
            out_channels=old_conv.out_channels,
            kernel_size=old_conv.kernel_size,
            stride=old_conv.stride,
            padding=old_conv.padding,
            bias=old_conv.bias is not None
        )

        with torch.no_grad():
            # Copy weights: L (first 3) and R (second 3) both get pretrained weights
            new_conv.weight[:, :3] = old_conv.weight.clone()
            new_conv.weight[:, 3:] = old_conv.weight.clone()

            if old_conv.bias is not None:
                new_conv.bias.copy_(old_conv.bias)

        self.backbone.stem[0] = new_conv  # Replace first conv layer

    def forward(self, x):
        x = self.backbone.forward_features(x)  # Extract features
        x = self.pool(x)
        x = self.flatten(x)
        return self.head(x)



# %% [markdown]
# The setup

# %%
model = ConvNeXt6DP6ch().to(DEVICE)  # Use the updated model class
optimizer = optim.Adam(model.parameters(), lr=LEARNING_RATE)
scheduler = optim.lr_scheduler.ReduceLROnPlateau(optimizer, patience=2)

# %% [markdown]
# Train func

# %%
def train(validate=True, resume_from_checkpoint=False):
    writer = SummaryWriter(log_dir=os.path.join(BASE_DIR, f"runs/S6ch-ConvNeXt6DP_batch{BATCH_ID}.{MOD_ID}"))
    scaler = GradScaler()
    best_val_loss = float('inf')
    epochs_no_improve = 0
    start_epoch = 0
    now = []
    time_per_epoch = []

    if resume_from_checkpoint:
        checkpoint = torch.load(MODEL_SAVE_PATH)
        model.load_state_dict(checkpoint['model_state_dict'])
        optimizer.load_state_dict(checkpoint['optimizer_state_dict'])
        scaler.load_state_dict(checkpoint['scaler_state_dict'])
        best_val_loss = checkpoint.get('best_val_loss', float('inf'))
        epochs_no_improve = checkpoint.get('epochs_no_improve', 0)
        start_epoch = checkpoint.get('epoch', 0)
        print(f"✅ Resumed from checkpoint at epoch {start_epoch}")
        now = [0.0] * (start_epoch + 1)

    if len(now) <= start_epoch:
        now.append(time.time())
    else:
        now[start_epoch] = time.time()

    for epoch in range(start_epoch, NUM_EPOCHS):
        print(f"\n� EPOCH : {epoch + 1}")
        model.train()
        train_loss, total_trans_rmse, total_rot_rmse = 0.0, 0.0, 0.0
        num_samples = 0
        pbar = tqdm(train_loader, desc=f"Epoch {epoch + 1}/{NUM_EPOCHS}", leave=False)

        for images, labels in pbar:
            images = images.to(DEVICE)
            labels = labels.to(DEVICE)
            batch_size = images.size(0)

            optimizer.zero_grad(set_to_none=True)
            with autocast(device_type="cuda"):
                outputs = model(images)
                loss = combined_loss(outputs, labels, TRANS_WEIGHT, ROTATION_WEIGHT, ANGULAR_WEIGHT)

            scaler.scale(loss).backward()
            scaler.step(optimizer)
            scaler.update()
            torch.cuda.synchronize()

            train_loss += loss.item() * batch_size
            trans_rmse, rot_rmse = compute_errors(outputs, labels)
            total_trans_rmse += trans_rmse * batch_size
            total_rot_rmse += rot_rmse * batch_size
            num_samples += batch_size

            pbar.set_postfix({"Loss": loss.item()})

        avg_train_loss = train_loss / num_samples
        avg_trans_rmse = total_trans_rmse / num_samples
        avg_rot_rmse = total_rot_rmse / num_samples

        now.append(time.time())
        time_per_epoch.append(int(now[epoch + 1] - now[epoch]))

        print(f"✅ Avg Training Loss: {avg_train_loss:.4f}")
        print(f"� RMSE - Trans: {avg_trans_rmse:.4f}, Rot: {avg_rot_rmse:.4f}")
        print(f"⏱️ Time per epoch: {time_per_epoch[epoch]}s")

        writer.add_scalar("Loss/Train", avg_train_loss, epoch)
        writer.add_scalar("RMSE/Train_Translation", avg_trans_rmse, epoch)
        writer.add_scalar("RMSE/Train_Rotation", avg_rot_rmse, epoch)

        if epoch != 0 and epoch % 5 == 0:
            parameters_to_prune = [
                (m, 'weight') for m in model.modules()
                if isinstance(m, (nn.Linear, nn.Conv2d)) and hasattr(m, 'weight')
            ]
            if parameters_to_prune:
                prune.global_unstructured(
                    parameters_to_prune,
                    pruning_method=prune.L1Unstructured,
                    amount=0.1
                )
                print(f"⚠️ Pruning applied at epoch {epoch}")

        if validate:
            model.eval()
            val_loss, val_trans_rmse, val_rot_rmse = 0.0, 0.0, 0.0
            val_samples = 0
            with torch.no_grad():
                for images, labels in val_loader:
                    images = images.to(DEVICE)
                    labels = labels.to(DEVICE)
                    batch_size = images.size(0)
                    with autocast(device_type="cuda"):
                        outputs = model(images)
                        loss = combined_loss(outputs, labels, TRANS_WEIGHT, ROTATION_WEIGHT, ANGULAR_WEIGHT)
                    torch.cuda.synchronize()
                    val_loss += loss.item() * batch_size
                    trans_rmse, rot_rmse = compute_errors(outputs, labels)
                    val_trans_rmse += trans_rmse * batch_size
                    val_rot_rmse += rot_rmse * batch_size
                    val_samples += batch_size

            avg_val_loss = val_loss / val_samples
            avg_val_trans_rmse = val_trans_rmse / val_samples
            avg_val_rot_rmse = val_rot_rmse / val_samples

            print(f"� Val Loss: {avg_val_loss:.4f} | RMSE Trans: {avg_val_trans_rmse:.4f}, Rot: {avg_val_rot_rmse:.4f}")

            writer.add_scalar("Loss/Val", avg_val_loss, epoch)
            writer.add_scalar("RMSE/Val_Translation", avg_val_trans_rmse, epoch)
            writer.add_scalar("RMSE/Val_Rotation", avg_val_rot_rmse, epoch)

            scheduler.step(avg_val_loss)

            if epoch != 0 and avg_val_loss < best_val_loss:
                best_val_loss = avg_val_loss
                epochs_no_improve = 0
                torch.save({
                    'model_state_dict': model.state_dict(),
                    'optimizer_state_dict': optimizer.state_dict(),
                    'scaler_state_dict': scaler.state_dict(),
                    'best_val_loss': best_val_loss,
                    'epochs_no_improve': epochs_no_improve,
                    'epoch': epoch + 1
                }, BEST_MODEL_PATH)
                print(f"� Best model saved to: {BEST_MODEL_PATH}")
            else:
                epochs_no_improve += 1
                print(f"⚠️ No improvement ({epochs_no_improve}/{PATIENCE})")
                if epochs_no_improve >= PATIENCE:
                    print("⏹️ Early stopping triggered")
                    break

        torch.save({
            'model_state_dict': model.state_dict(),
            'optimizer_state_dict': optimizer.state_dict(),
            'scaler_state_dict': scaler.state_dict(),
            'best_val_loss': best_val_loss,
            'epochs_no_improve': epochs_no_improve,
            'epoch': epoch + 1
        }, MODEL_SAVE_PATH)
        print(f"� Checkpoint saved to: {MODEL_SAVE_PATH}")

    writer.close()
    return time_per_epoch


# %% [markdown]
# Actually training

# %%
time_per_epoch = train(validate=True,resume_from_checkpoint=False)

# %% [markdown]
# Update

# %%
# gv['mod_id'] += 1
# with open("GlobVar.json", "w") as file:
#     json.dump(gv, file, indent=4)
# print("mod_id updated in GlobVar.json")

# %%
CLEAN_MODEL_PATH = os.path.join(BASE_DIR, f"model/CLEAN-S6ch-ConvNeXt6DP{BATCH_ID}.{MOD_ID}.pth")
for name, module in model.named_modules():
    if isinstance(module, torch.nn.Conv2d) or isinstance(module, torch.nn.Linear):
        if hasattr(module, "weight_orig"):
            prune.remove(module, "weight")
torch.save(model.state_dict(), CLEAN_MODEL_PATH)

# %% [markdown]
# Test func

# %%
def test_model(model, loader, mode='Test', use_amp=False):
    model.eval()
    inference_times = []
    total_loss = 0.0
    total_trans_rmse = 0.0
    total_rot_rmse = 0.0
    num_samples = 0
    all_preds, all_gts = [], []

    with torch.no_grad():
        for images, labels in tqdm(loader, desc=f"Running {mode}"):
            images = images.to(DEVICE)
            labels = labels.to(DEVICE)
            batch_size = images.size(0)

            torch.cuda.synchronize()
            start_time = time.time()

            if use_amp:
                with autocast(device_type="cuda"):
                    outputs = model(images)
                    loss = combined_loss(outputs, labels, TRANS_WEIGHT, ROTATION_WEIGHT, ANGULAR_WEIGHT)
            else:
                outputs = model(images)
                loss = combined_loss(outputs, labels, TRANS_WEIGHT, ROTATION_WEIGHT, ANGULAR_WEIGHT)

            torch.cuda.synchronize()
            inference_time = (time.time() - start_time) * 1000 / batch_size  # ms per image
            inference_times.append(inference_time)

            total_loss += loss.item() * batch_size
            trans_rmse, rot_rmse = compute_errors(outputs, labels)
            total_trans_rmse += trans_rmse * batch_size
            total_rot_rmse += rot_rmse * batch_size
            num_samples += batch_size

            all_preds.append(outputs.cpu().numpy())
            all_gts.append(labels.cpu().numpy())

    avg_loss = total_loss / num_samples
    avg_trans_rmse = total_trans_rmse / num_samples
    avg_rot_rmse = total_rot_rmse / num_samples
    avg_inference_time = np.mean(inference_times)

    return avg_loss, avg_trans_rmse, avg_rot_rmse, np.concatenate(all_preds), np.concatenate(all_gts), avg_inference_time


# %% [markdown]
# Actually testing

# %%
model.load_state_dict(torch.load(CLEAN_MODEL_PATH))
model.to(DEVICE)

test_avg_loss, test_avg_trans_rmse, test_avg_rot_rmse, test_preds, test_gts, test_avg_inference_time = test_model(
    model, test_loader, mode='Test', use_amp=True
)

print("\n📊 Test Summary")
print(f"🔁 Average Loss       : {test_avg_loss:.4f}")
print(f"📐 Translation RMSE   : {test_avg_trans_rmse:.4f}")
print(f"📐 Rotation RMSE      : {test_avg_rot_rmse:.4f}")
print(f"⚡ Inference Time (ms) : {test_avg_inference_time:.2f} ms/image")

# %% [markdown]
# Val func

# %%
def validate_model(model_path=None):
    inference_times = []
    if model_path:
        if not os.path.exists(model_path):
            raise ValueError(f"Model path {model_path} does not exist.")
        model.load_state_dict(torch.load(model_path))
    model.to(DEVICE)
    model.eval()
    total_loss, total_trans_rmse, total_rot_rmse = 0.0, 0.0, 0.0
    num_samples = 0
    all_preds, all_gts = [], []

    with torch.no_grad():
        for images, labels in tqdm(val_loader, desc="Running Validation"):
            images = images.to(DEVICE)
            labels = labels.to(DEVICE)
            batch_size = images.size(0)

            torch.cuda.synchronize()
            start_time = time.time()

            with autocast(device_type="cuda"):
                outputs = model(images)
                loss = combined_loss(outputs, labels, TRANS_WEIGHT, ROTATION_WEIGHT, ANGULAR_WEIGHT)

            torch.cuda.synchronize()
            inference_time = (time.time() - start_time) * 1000 / batch_size  # ms per image
            inference_times.append(inference_time)

            total_loss += loss.item() * batch_size
            trans_rmse, rot_rmse = compute_errors(outputs, labels)
            total_rot_rmse += rot_rmse * batch_size
            total_trans_rmse += trans_rmse * batch_size
            num_samples += batch_size

            all_preds.append(outputs.cpu().numpy())
            all_gts.append(labels.cpu().numpy())

    avg_inference_time = np.mean(inference_times)
    avg_loss = total_loss / num_samples
    avg_trans_rmse = total_trans_rmse / num_samples
    avg_rot_rmse = total_rot_rmse / num_samples

    return avg_loss, avg_trans_rmse, avg_rot_rmse, np.concatenate(all_preds), np.concatenate(all_gts), avg_inference_time


# %% [markdown]
# Actually validating

# %%
val_avg_loss, val_avg_trans_rmse, val_avg_rot_rmse, val_preds, val_gts, val_avg_inference_time = validate_model(CLEAN_MODEL_PATH)

# %%
torch.cuda.empty_cache()

# %% [markdown]
# Calculating 

# %%
test_translation_rmse_cm = calculate_translation_rmse(test_preds, test_gts)
val_translation_rmse_cm = calculate_translation_rmse(val_preds, val_gts)
test_rot_accuracy = rotation_error_deg_from_6d(torch.tensor(test_preds[:, 3:]), torch.tensor(test_gts[:, 3:]))
val_rot_accuracy = rotation_error_deg_from_6d(torch.tensor(val_preds[:, 3:]), torch.tensor(val_gts[:, 3:]))

train_trans_stats = get_dataset_stats(train_loader)
val_trans_stats = get_dataset_stats(val_loader)
test_trans_stats = get_dataset_stats(test_loader)

test_range_cm = (test_trans_stats['max'].mean() - test_trans_stats['min'].mean()) * 100
val_range_cm = (val_trans_stats['max'].mean() - val_trans_stats['min'].mean()) * 100

test_trans_accuracy_pct = translation_accuracy_percentage(test_translation_rmse_cm, test_range_cm).item()
val_trans_accuracy_pct = translation_accuracy_percentage(val_translation_rmse_cm, val_range_cm).item()

test_rot_accuracy_pct = 100*(1-(test_rot_accuracy.item()/360))
val_rot_accuracy_pct = 100*(1-(val_rot_accuracy.item()/360))

# %% [markdown]
# Write MD

# %%
eval_path = os.path.join(BASE_DIR, f"model/S6ch-ConvNeXt6DP_batch{BATCH_ID}.{MOD_ID}.md")
eval_content = f"""# Evaluation Results - Batch {BATCH_ID} - Model {MOD_ID}

## Training Configuration
- Batch Size: {BATCH_SIZE}
- Epochs: {NUM_EPOCHS}
- Learning Rate: {LEARNING_RATE}
- Translation Weight : {TRANS_WEIGHT}
- Rotation Weight : {ROTATION_WEIGHT}
- Angular Weight : {ANGULAR_WEIGHT}
- Patience : {PATIENCE}
- Image Size: {IMG_SIZE}
- Device: {DEVICE}
- Optimizer : Adam

## Model Architecture
- Backbone: Using ConvNeXtV2 Nano @ {IMG_SIZE} Image split and stacked into a 6Ch image
- Head: Linear(620 -> 512 -> 9)

## Evaluation Metrics

### Test Set
- Average Loss: {test_avg_loss:.4f}
- Translation RMSE: {test_avg_trans_rmse:.4f}
- Translation Accuracy: {test_translation_rmse_cm:.2f} cm
- Translation Accuracy %: {test_trans_accuracy_pct:.2f}%
- Rotation RMSE: {test_avg_rot_rmse :.4f}
- Rotation Accuracy: {test_rot_accuracy:.2f}°
- Rotation Accuracy % : {test_rot_accuracy_pct:.2f} %
- Inference Speed: {test_avg_inference_time:.2f} ms/frame

### Validation Set
- Average Loss: {val_avg_loss:.4f}
- Translation RMSE: {val_avg_trans_rmse :.4f}
- Translation Accuracy: {val_translation_rmse_cm:.2f} cm
- Translation Accuracy %: {val_trans_accuracy_pct:.2f}%
- Rotation RMSE: {val_avg_rot_rmse :.4f}
- Rotation Accuracy: {val_rot_accuracy:.2f}°
- Rotation Accuracy % : {val_rot_accuracy_pct:.2f} %
- Inference Speed: {val_avg_inference_time:.2f} ms/frame

## Dataset Statistics
### Training Set
- Translation range: [{train_trans_stats['min'].mean():.2f}, {train_trans_stats['max'].mean():.2f}] m

### Validation Set
- Translation range: [{val_trans_stats['min'].mean():.2f}, {val_trans_stats['max'].mean():.2f}] m

### Test Set
- Translation range: [{test_trans_stats['min'].mean():.2f}, {test_trans_stats['max'].mean():.2f}] m

## File Locations
- Dataset Directory: {DATASET_DIR}
- Model Save Path: {MODEL_SAVE_PATH}
"""

with open(eval_path, 'w') as f:
    f.write(eval_content)
print(f"Evaluation report saved to: {eval_path}")

# %%
csv_path = os.path.join(BASE_DIR, "model/eval_results.csv")
write_header = not os.path.exists(csv_path)

csv_data = {
    'timestamp': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    'dataset_id': BATCH_ID,
    'model_id': MOD_ID,
    'batch_size': BATCH_SIZE,
    'epochs': NUM_EPOCHS,
    'learning_rate': LEARNING_RATE,
    'translation_weight':TRANS_WEIGHT,
    'rotation_weight' : ROTATION_WEIGHT,
    'angular_weight' : ANGULAR_WEIGHT,
    'patience' : PATIENCE,
    'test_loss': test_avg_loss,
    'test_translation_rmse': test_avg_trans_rmse,
    'test_translation_accuracy_pct': test_trans_accuracy_pct,
    'test_rotation_rmse': test_avg_rot_rmse ,
    'test_rotation_accuracy_pct':test_rot_accuracy_pct,
    'test_inference_time_ms': test_avg_inference_time,
    'validation_loss': val_avg_loss,
    'validation_translation_rmse': val_avg_trans_rmse ,
    'validation_translation_accuracy_pct': val_trans_accuracy_pct,
    'validation_rotation_rmse': val_avg_rot_rmse ,
    'validation_rotation_accuracy_pct':val_rot_accuracy_pct,
    'validation_inference_time_ms': val_avg_inference_time,
    'model_path': MODEL_SAVE_PATH,
    'eval_path' : eval_path
}

with open(csv_path, 'a', newline='') as csvfile:
    writer = csv.DictWriter(csvfile, fieldnames=csv_data.keys())
    if write_header:
        writer.writeheader()
    writer.writerow(csv_data)
print(f"Results appended to CSV: {csv_path}")

# %%
def safe(val):
    if isinstance(val, (np.floating, np.integer)):
        return val.item()
    return val

def log_eval_to_xlsx(
    variant="6ch",
    head_arch="640 → 512 → 9",
    param_count=332288,
    notes="Checking on 51, waiting evaluation of both",
    xlsx_path=os.path.join(BASE_DIR, "model/EVAL.xlsx")
):
    headers = [
        "ID", "IMG_SIZE", "VARIANT", "HEAD_ARCH", "PARAM_COUNT", "BATCH_SIZE", "EPC", "DTS_ID", "DTS_LEN", "LR_RT",
        "TRAIN_TIME", "VRAM_USG", "VAL_LOSS", "TS_LOSS",
        "VAL_TRANS_RSME", "TS_TRANS_RMSE", "VAL_ROT_RSME", "TS_ROT_RMSE",
        "VAL_TRANS_ACC", "TS_TRANS_ACC", "VAL_ROT_ACC", "TS_ROT_ACC",
        "VAL_INF_MS", "TS_INF_MS", "Notes"
    ]

    model_id_str = f"{BATCH_ID}{MOD_ID}"
    dataset_len = len(train_loader.dataset) + len(val_loader.dataset) + len(test_loader.dataset)

    row = [
        int(model_id_str), IMG_SIZE, variant, head_arch, param_count, BATCH_SIZE, NUM_EPOCHS,
        BATCH_ID, dataset_len, str(LEARNING_RATE),  # Still keep LR as string
        round(sum(time_per_epoch)/60,1), 7.5,  # TRAIN_TIME, VRAM_USG

        round(val_avg_loss, 4), round(test_avg_loss, 4),
        round(val_avg_trans_rmse , 4),
        round(test_avg_trans_rmse, 4),
        round(val_avg_rot_rmse , 4),
        round(test_avg_rot_rmse , 4),

        round(val_trans_accuracy_pct, 2), round(test_trans_accuracy_pct, 2),
        round(val_rot_accuracy_pct, 2), round(test_rot_accuracy_pct, 2),

        round(val_avg_inference_time, 2), round(test_avg_inference_time, 2),

        notes
    ]


    row = [safe(x) for x in row]

    # Load or create workbook
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "ConvNeXt V2 Nano"
        ws.append(headers)

    ws.append(row)
    wb.save(xlsx_path)
    print(f"✅ Evaluation logged to Excel: {xlsx_path}")


log_eval_to_xlsx()


